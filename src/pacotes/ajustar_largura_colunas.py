from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tkinter import messagebox

def ajustar_largura_colunas(arquivo, max_linhas=100):
    """
    Ajusta a largura das colunas no arquivo Excel de forma eficiente, 
    combinando larguras predefinidas para colunas específicas e ajuste dinâmico para outras.
    
    Args:
        arquivo (str): Caminho do arquivo Excel.
        max_linhas (int): Número máximo de linhas para estimar a largura das colunas dinamicamente.
    """
    try:
        wb = load_workbook(arquivo)
        ws = wb.active

        # Larguras predefinidas para colunas específicas
        largura_colunas_predefinidas = {
            "A": 5,   # UT
            "B": 5,   # Faixa
            "C": 5,   # Placa
            "D": 20,  # Nome Vulgar
            "E": 25,  # Nome Cientifico
            "F": 5,   # CAP
            "G": 5,   # ALT
            "H": 5,   # QF
            "I": 5,   # X
            "J": 5,   # Y
            "K": 10,  # DAP
            "L": 15,  # Volume_m3
            "M": 15,  # Latitude
            "N": 15,  # Longitude
            "O": 10,  # DM
            "P": 15,  # Observacoes
            "Q": 15   # Categoria
        }

        # Ajuste da largura das colunas
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)  # Obtém a letra da coluna
            max_length = 0

            # Caso exista largura predefinida, aplica diretamente
            if col_letter in largura_colunas_predefinidas:
                ws.column_dimensions[col_letter].width = largura_colunas_predefinidas[col_letter]
                continue

            # Caso contrário, calcula a largura dinamicamente
            for i, cell in enumerate(col):
                if i >= max_linhas:  # Limita o cálculo às primeiras `max_linhas` linhas
                    break
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # Define largura ajustada ou largura mínima (10 por padrão)
            adjusted_width = max(max_length + 2, 10)
            ws.column_dimensions[col_letter].width = adjusted_width

        # Salva o arquivo com as larguras ajustadas
        wb.save(arquivo)
        print(f"Largura das colunas ajustada com sucesso em: {arquivo}")
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao ajustar as larguras das colunas: {e}")
